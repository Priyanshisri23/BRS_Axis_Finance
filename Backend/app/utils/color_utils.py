import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def color_excel(brs_output_file,t_1_date):
    # Load and validate BRS file
    brs_workbook = openpyxl.load_workbook(brs_output_file)
    sheet=brs_workbook[t_1_date]

    # Apply cell style for the data rows (yellow background for values)
    key_fill = PatternFill(start_color="BA0465", end_color="BA0465", fill_type="solid")
    key_font = Font(bold=True, color="FFFFFF")
    value_font = Font(bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    # Define border styles
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    # Apply the styles to cells A2 to A10
    for row in range(2, 13):  # Rows 2 to 10
        key_cell = sheet[f'B{row}']
        value_cell = sheet[f'C{row}']

        
        # Apply the styles
        key_cell.fill = key_fill
        key_cell.font = key_font
        key_cell.alignment = alignment
        key_cell.border = thin_border
        
        # value_cell.fill = value_fill
        value_cell.font = value_font
        value_cell.alignment = alignment
        value_cell.border = thin_border

        for col in range(2, sheet.max_column + 1):  # Loop through all columns in row 13
            cell = sheet.cell(row=14, column=col)
            cell.fill = key_fill
            cell.font = value_font
            cell.alignment = alignment
            cell.border = thin_border

        # Adjust column width for better readability
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 40

        brs_workbook.save(brs_output_file)